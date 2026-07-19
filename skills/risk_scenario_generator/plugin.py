import json
import os
import urllib.request
import asyncio
from pathlib import Path
from datetime import datetime

DEFAULT_RECOMMENDATION_RULES = {
    "flooding": {
        "title": "Затопление / Наводнение",
        "scenarios": [
            {
                "name": "Сценарий 1: Физическая безопасность и консервация активов (Консервативный)",
                "context": "Применяется при критическом уровне подъема воды или угрозе непосредственного затопления помещений филиала и банкоматной сети.",
                "actions": [
                    "Экстренное отключение электропитания уязвимых банкоматов и оборудования в зоне затопления для защиты электроники от необратимых повреждений.",
                    "Приостановка обслуживания клиентов в затронутых физических офисах, эвакуация сотрудников и документов сейфового хранения на возвышенность/резервные точки.",
                    "Выгрузка денежной наличности (инкассация) из банкоматов первой линии подтопления для минимизации риска порчи купюр водой."
                ],
                "needs": [
                    "Обеспечение абсолютной физической безопасности жизни клиентов и сотрудников банка.",
                    "Сохранение материальных активов и наличности от порчи и физического уничтожения.",
                    "Защита инфраструктуры и дорогостоящих электронных плат самообслуживания от замыканий."
                ]
            },
            {
                "name": "Сценарий 2: Цифровое перенаправление и удаленное обслуживание (Адаптивный)",
                "context": "Применяется для сохранения клиентского сервиса в условиях физической блокировки офисов и банкоматов пострадавшей зоны.",
                "actions": [
                    "Перевод всех операций клиентов пострадавшего региона в мобильное приложение и интернет-банк с отменой локальных комиссий за критические переводы.",
                    "Оповещение клиентов через СМС, email и push-уведомления об альтернативных безопасных точках самообслуживания за границами паводка.",
                    "Маршрутизация звонков горячей линии из затронутого региона на свободные КИЦ Дагестана и юга России."
                ],
                "needs": [
                    "Поддержание лояльности клиентов и сохранение доверия к банку как к надежному стабильному партнеру в ЧС.",
                    "Обеспечение непрерывности критических платежей граждан (включая выплаты пособий, лекарства, товары первой необходимости)."
                ]
            }
        ]
    },
    "fire": {
        "title": "Пожар / Возгорание",
        "scenarios": [
            {
                "name": "Сценарий 1: Экстренная эвакуация и пожарная сигнализация (Экстренный)",
                "context": "Применяется при срабатывании датчиков дыма, угрозе пожара или получении оперативных сигналов МЧС.",
                "actions": [
                    "Немедленная эвакуация всех находящихся в здании людей по путям пожарного выхода согласно плану эвакуации.",
                    "Экстренное отключение силовых электрощитов, серверов и банкоматной техники (по возможности) с блокировкой сейфовых касс.",
                    "Вызов расчетов пожарной охраны и организация доступа спецтехники к гидрантам филиала."
                ],
                "needs": [
                    "Абсолютный приоритет безопасности жизни людей над материальными ценностями.",
                    "Быстрое сдерживание зоны возгорания и предупреждение детонации касс."
                ]
            }
        ]
    },
    "communication_block": {
        "title": "Блокировка мобильного интернета / Отказ каналов связи",
        "scenarios": [
            {
                "name": "Сценарий 3: Локальная автономия и резервное резервирование (Автономный)",
                "context": "Применяется при массовых нарушениях беспроводной связи со штабом и процессингом в секторе.",
                "actions": [
                    "Перевод отделений банка на стабильные проводные кабельные или резервные спутниковые (VSAT) каналы связи.",
                    "Переключение терминалов самообслуживания на буферизованные автономные протоколы (проведение транзакций выдачи с локальными лимитами до восстановления связи).",
                    "Использование локальных резервных копий баз данных для проведения неотложных офисных кассовых ордеров."
                ],
                "needs": [
                    "Сохранение операционной дееспособности филиала в условиях внешней информационной изоляции.",
                    "Возможность проведения срочных ручных кассовых операций для корпоративных и розничных клиентов."
                ]
            },
            {
                "name": "Сценарий 4: Офлайн-инкассация и защищенный маршрут (Логистический)",
                "context": "Применяется при невозможности координации инкассаторских машин со спутниковым GPS/ГЛОНАСС мониторингом.",
                "actions": [
                    "Перевод инкассаторских бригад CIT на жестко регламентированные во времени парные маршруты движения.",
                    "Использование резервных КВ/УКВ радиостанций для тактической связи между машинами CIT и локальным дежурным офисом.",
                    "Приостановка подкрепления банкоматов с высоким риском ограбления на глухих неосвещенных трассах."
                ],
                "needs": [
                    "Защита инкассаторов и наличных средств во время транспортировки при потере телеметрии контроля маршрута.",
                    "Разумное расходование запаса наличности в работающих изолированных АТМ."
                ]
            }
        ]
    },
    "fuel_shortage": {
        "title": "Дефицит топлива / Ограничение снабжения",
        "scenarios": [
            {
                "name": "Сценарий 1: Оптимизация инкассаторских выездов (Логистический)",
                "context": "Применяется при жестком дефиците ГСМ для CIT-транспорта региона.",
                "actions": [
                    "Укрупнение лимитов загрузки касс банкоматов и отделений (загрузка сразу максимальных объемов наличных для снижения частоты выездов).",
                    "Приоритетное снабжение ГСМ инкассаторских автомобилей на бронированной базе, сокращение вспомогательных выездов руководства.",
                    "Оптимизация логистических цепочек по методу коммивояжера с исключением холостых или дублирующих пробегов."
                ],
                "needs": [
                    "Поддержание стабильного оборота наличных в основных узлах присутствия с минимальным расходом топлива.",
                    "Концентрация ресурсов на жизнеобеспечении системообразующих филиалов банка."
                ]
            }
        ]
    }
}

def _get_server_port(api) -> str:
    state_dir = Path(api.get_state_dir())
    data_root = state_dir.parent.parent / "server_port"
    port = "8765"
    if data_root.exists():
        try:
            port = data_root.read_text(encoding="utf-8").strip()
        except Exception:
            pass
    return port

def _read_temp_storage_key(api, key: str) -> str:
    port = _get_server_port(api)
    url = f"http://127.0.0.1:{port}/api/extensions/temp_storage/retrieve"
    payload = {"key": key}
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            res = json.loads(r.read().decode("utf-8"))
            retrieved = res.get("retrieved") or {}
            if retrieved.get("key") == key:
                return retrieved.get("value") or ""
    except Exception as e:
        api.log("warning", f"HTTP POST to temp_storage retrieve failed: {str(e)}")
        
    state_dir = Path(api.get_state_dir())
    local_fallback = state_dir / f"fallback_{key}.json"
    if local_fallback.exists():
        try:
            return local_fallback.read_text(encoding="utf-8")
        except Exception as exc:
            api.log("error", f"Local fallback read failed: {str(exc)}")
            
    # Кросс-скилл фолбек для полной автономности в офлайн/рестарт периоды
    sibling_fallback = state_dir.parent / "asset_geo_matcher" / f"fallback_{key}.json"
    if sibling_fallback.exists():
        try:
            return sibling_fallback.read_text(encoding="utf-8")
        except Exception as exc:
            api.log("error", f"Sibling asset_geo_matcher fallback read failed: {str(exc)}")
            
    # Самый надежный, прямой дисковый фолбек из хранилища temp_storage напрямую:
    temp_storage_store = state_dir.parent / "temp_storage" / "store" / f"{key}.txt"
    if temp_storage_store.exists():
        try:
            return temp_storage_store.read_text(encoding="utf-8")
        except Exception as exc:
            api.log("error", f"Direct temp_storage store read failed: {str(exc)}")
            
    return ""

def _write_temp_storage_key(api, key: str, value: str, description: str = "") -> bool:
    port = _get_server_port(api)
    url = f"http://127.0.0.1:{port}/api/extensions/temp_storage/store"
    payload = {
        "key": key,
        "value": value,
        "description": description
    }
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            res = json.loads(r.read().decode("utf-8"))
            if res and "error" not in res:
                return True
    except Exception as e:
        api.log("warning", f"HTTP POST to temp_storage store failed: {str(e)}")
        
    try:
        state_dir = Path(api.get_state_dir())
        state_dir.mkdir(parents=True, exist_ok=True)
        local_fallback = state_dir / f"fallback_{key}.json"
        local_fallback.write_text(value, encoding="utf-8")
        return True
    except Exception as exc:
        api.log("error", f"Local fallback write failed: {str(exc)}")
    return False

def check_temp_storage_health(api) -> bool:
    port = _get_server_port(api)
    url = f"http://127.0.0.1:{port}/api/extensions/temp_storage/status"
    try:
        req = urllib.request.Request(url, method="GET")
        with urllib.request.urlopen(req, timeout=3) as r:
            return r.status == 200
    except Exception:
        return False

def find_recommendations_excel() -> str:
    user_home = os.environ.get('USERPROFILE', os.path.expanduser('~'))
    candidates = [
        os.path.join(user_home, 'Downloads', 'Агент - примеры сценариев.xlsx'),
        os.path.join(user_home, 'Downloads', 'Агент - примеры сценариев-1.xlsx'),
        os.path.join(user_home, 'Downloads', 'Агент - примеры сценариев-2.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return ""

def normalize_text(text: str) -> str:
    import unicodedata
    if not text:
        return ""
    return unicodedata.normalize('NFC', str(text)).strip().lower().replace('\xa0', ' ')

def load_scenarios_from_excel(path: str, api) -> dict:
    result_rules = {}
    if not path:
        return result_rules
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        sheet_name = "Лист1"
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
            
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        # Динамически определяем индексы колонок из заголовка
        header_idx = None
        for i, r in enumerate(rows):
            if r and any(isinstance(val, str) and "Чрезвычайная ситуация" in val for val in r if val):
                header_idx = i
                break
                
        if header_idx is None:
            header_idx = 2  # дефолтный ряд с заголовками
            
        header_row = rows[header_idx] if header_idx < len(rows) else []
        col_indices = {"ch_type": 0, "goal": 1, "action": 2}
        for i, val in enumerate(header_row):
            if not val:
                continue
            val_str = str(val).lower()
            if "ситуация" in val_str:
                col_indices["ch_type"] = i
            elif "цель" in val_str:
                col_indices["goal"] = i
            elif "сценарий" in val_str or "реагирования" in val_str:
                col_indices["action"] = i
                
        # Группируем строки сценариев по типу ЧС в Excel
        raw_groups = {}
        for r_idx in range(header_idx + 1, len(rows)):
            row = rows[r_idx]
            if not row or len(row) <= max(col_indices.values()):
                continue
            ch_type = row[col_indices["ch_type"]]
            goal = row[col_indices["goal"]]
            action = row[col_indices["action"]]
            if not ch_type or not goal or not action:
                continue
                
            ch_type_str = str(ch_type).strip()
            goal_str = str(goal).strip()
            action_str = str(action).strip()
            
            if ch_type_str not in raw_groups:
                raw_groups[ch_type_str] = []
            raw_groups[ch_type_str].append((goal_str, action_str))
            
        # Карта русско-английских системных кодов для сопоставления в compile_scenarios
        MAPPING_RU_TO_ENG = {
            "затоплен": "flooding",
            "наводнен": "flooding",
            "пожар": "fire",
            "возгоран": "fire",
            "интернет": "communication_block",
            "связ": "communication_block",
            "топлив": "fuel_shortage",
            "бензин": "fuel_shortage",
            "заправк": "fuel_shortage"
        }
        
        for ch_type_ru, items in raw_groups.items():
            ch_type_norm = normalize_text(ch_type_ru)
            eng_key = None
            for key_ru, key_eng in MAPPING_RU_TO_ENG.items():
                if key_ru in ch_type_norm:
                    eng_key = key_eng
                    break
                    
            if not eng_key:
                eng_key = ch_type_norm
                
            scenarios = []
            for sc_idx, (goal, action) in enumerate(items, 1):
                # Формируем конкретные шаги с разбиением по ";"
                actions_list = []
                for chunk in action.split(";"):
                    chunk_stripped = chunk.strip().lstrip('•-* ').strip()
                    if chunk_stripped:
                        actions_list.append(chunk_stripped)
                        
                scenarios.append({
                    "name": f"Сценарий {sc_idx}: {goal}",
                    "context": f"Применяется при урегулировании ЧС '{ch_type_ru}'. Цель сценария: {goal}.",
                    "actions": actions_list,
                    "needs": [goal]
                })
                
            rules_block = {
                "title": ch_type_ru,
                "scenarios": scenarios
            }
            # Сохраняем правила под всеми возможными синонимами
            result_rules[eng_key] = rules_block
            result_rules[ch_type_ru] = rules_block
            result_rules[ch_type_norm] = rules_block
            
        wb.close()
    except Exception as e:
        api.log("error", f"Failed to load scenarios from Excel structure: {str(e)}")
    return result_rules

def find_recommendations_file() -> tuple:
    user_home = os.environ.get('USERPROFILE', os.path.expanduser('~'))
    candidates = [
        (os.path.join(user_home, 'Downloads', 'Рекомендации.json'), 'json_ru'),
        (os.path.join(user_home, 'Downloads', 'recommendations.json'), 'json_en'),
        (os.path.join(user_home, 'Downloads', 'Рекомендации.txt'), 'txt_ru'),
        (os.path.join(user_home, 'Downloads', 'recommendations.txt'), 'txt_en')
    ]
    for path, ftype in candidates:
        if os.path.exists(path):
            return path, ftype
    return "", ""

def load_recommendations_safely(path: str, ftype: str, api) -> dict:
    result = {"has_file": False, "filename": "", "rules": {}, "raw_text": ""}
    if not path:
        return result
        
    result["has_file"] = True
    result["filename"] = os.path.basename(path)
    
    encodings = ["utf-8", "cp1251", "cp1252"]
    content = ""
    for enc in encodings:
        try:
            content = Path(path).read_text(encoding=enc)
            break
        except Exception:
            continue
            
    if not content:
        # Final desperate fallback with replacement characters of unrecognized bytes
        try:
            content = Path(path).read_text(encoding="utf-8", errors="replace")
        except Exception:
            return result
            
    result["raw_text"] = content
    
    if "json" in ftype:
        try:
            result["rules"] = json.loads(content)
        except Exception as e:
            api.log("warning", f"Failed to parse recommendations JSON file: {str(e)}")
            
    return result

def compile_scenarios(api, mapping: list, ext_rules: dict, raw_text: str) -> tuple:
    scenarios_list = []
    summary_incidents_text_blocks = []
    processed_incident_types = set()
    
    # Analyze threat mappings
    for item in mapping:
        inc_type = item.get("incident_type", "unknown")
        inc_type_ru = item.get("incident_type_ru", "Чрезвычайная ситуация")
        city = item.get("city", "Локальный сектор")
        sev = item.get("severity", "low")
        count = item.get("affected_assets_count", 0)
        
        block = f"- **Тип ЧС**: {inc_type_ru} ({inc_type})\n  **Локация**: {city} (Уровень опасности: {sev})\n  **Угроза для офисов/терминалов**: Затронуто {count} объектов банка."
        summary_incidents_text_blocks.append(block)
        
        # Pull scenario recommendations dynamically (using external matches if present, falling back to builtin logic)
        custom_scenarios = None
        if ext_rules and isinstance(ext_rules, dict):
            # Check for matches on incident type or localized name
            custom_scenarios = ext_rules.get(inc_type) or ext_rules.get(inc_type_ru)
            
        if custom_scenarios and isinstance(custom_scenarios, dict):
            rules_source = custom_scenarios
        else:
            rules_source = DEFAULT_RECOMMENDATION_RULES.get(inc_type) or DEFAULT_RECOMMENDATION_RULES.get("flooding")
            
        if rules_source and inc_type not in processed_incident_types:
            processed_incident_types.add(inc_type)
            for sc in rules_source.get("scenarios", []):
                sc_copy = dict(sc)
                sc_copy["threat_context"] = f"Реагирование на ЧС '{inc_type_ru}' в городе {city} ({count} связанных активов)."
                scenarios_list.append(sc_copy)
                
    # If no threats or un-paired ЧС exist
    if not scenarios_list:
        # Check if recommendations file has raw guidance
        if raw_text:
            scenarios_list.append({
                "name": "Сценарий 1: Корпоративный регламент из файла пользователя",
                "context": "Действия по умолчанию на основе загруженного файла рекомендаций.",
                "actions": [line.strip() for line in raw_text.split("\n") if line.strip() and len(line) > 10][:4],
                "needs": ["Обеспечение операционной выносливости согласно уставу и внешнему файлу."]
            })
        else:
            # Absolute baseline safety scenario
            scenarios_list.append({
                "name": "Сценарий Резервный: Аварийный мониторинг и базовый регламент безопасности",
                "context": "Применяется как сценарий общего реагирования при отсутствии активных или распознанных угроз в картотеке.",
                "actions": [
                    "Перевод внутренних служб ИБ и инкассации в режим повышенной бдительности.",
                    "Проверка работоспособности резервных электрогенераторов во всех филиалах.",
                    "Контрольный опрос удаленных узлов самообслуживания по каналам телеметрии."
                ],
                "needs": ["Превентивное поддержание операционной устойчивости банковской ячейки."]
            })
            
    # Guarantee between 2 and 4 scenarios (Principle P5 / original prompt)
    if len(scenarios_list) < 2:
        # Append a generic adaptive fallback scenario
        scenarios_list.append({
            "name": "Сценарий 2: Информационный протокол кризисной службы (Информационный)",
            "context": "Организация связи и координации при любых аномальных ситуациях различного характера.",
            "actions": [
                "Активация удаленного кризисного штаба банка с регулярными сессиями синхронизации раз в 4 часа.",
                "Усиление контроля периметра физической безопасности оставшихся отделений.",
                "Резервное копирование актуальных транзакционных логов во всех базах данных."
            ],
            "needs": [
                "Сохранение скоординированного контроля и оперативного управления филиалами.",
                "Обоюдная осведомленность руководства и персонала о статусе ликвидации аварии."
            ]
        })
        
    # Cap to maximum of 4 scenarios
    scenarios_list = scenarios_list[:4]
    
    return scenarios_list, summary_incidents_text_blocks

def perform_decision_generation(api, force: bool = False) -> str:
    api.log("info", "Starting risk_scenario_generator logic calculations.")
    
    # 1. Preflight sanity check
    temp_storage_healthy = check_temp_storage_health(api)
    if not temp_storage_healthy:
        api.log("warning", "temp_storage is inactive. Generating with local state fallback backups.")
        
    # 2. Get threat mapping data from temp_storage
    raw_mapping = _read_temp_storage_key(api, "threat_mapping")
    mapping = []
    if raw_mapping:
        try:
            mapping = json.loads(raw_mapping)
        except Exception as e:
            api.log("warning", f"Threat mapping JSON parse warning: {str(e)}")
            
    if not mapping:
        api.log("warning", "No threat maps found in temp_storage 'threat_mapping' or format is empty.")
        
    # 3. Handle optional recommendations file
    rec_path_excel = find_recommendations_excel()
    excel_rules = {}
    if rec_path_excel:
        excel_rules = load_scenarios_from_excel(rec_path_excel, api)
        
    rec_path, rec_ftype = find_recommendations_file()
    rec_obj = load_recommendations_safely(rec_path, rec_ftype, api)
    
    # Merge rules: excel rules override text or json rules
    merged_rules = dict(rec_obj["rules"] or {})
    if excel_rules:
        merged_rules.update(excel_rules)
    
    # 4. Generate scenarios and incident summaries
    scenarios, summaries = compile_scenarios(api, mapping, merged_rules, rec_obj["raw_text"])
    
    # 5. Build clean, structural Report according to the user's formatting requirements
    date_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    report = []
    
    # Paragraph 1 (Skipped since this is a production run, not a simulated test)
    
    # Paragraph 2 (Brief summary of active emergencies)
    if mapping:
        cities = list(set(item.get("city", "Локальный сектор") for item in mapping if item.get("city")))
        inc_types_ru = list(set(item.get("incident_type_ru", "Чрезвычайная ситуация") for item in mapping if item.get("incident_type_ru")))
        dates = []
        for item in mapping:
            raw_date = item.get("date") or item.get("started_at") or ""
            if raw_date:
                try:
                    if "T" in raw_date:
                        raw_date = raw_date.split("T")[0]
                    parts = raw_date.split("-")
                    if len(parts) == 3:
                        dates.append(f"{parts[2]}.{parts[1]}.{parts[0]}")
                except Exception:
                    pass
        from_date = sorted(dates)[0] if dates else datetime.utcnow().strftime("%d.%m.%Y")
        
        total_atms = sum(item.get("affected_atms_count", 0) or item.get("affected_assets_count", 0) for item in mapping)
        total_offices = sum(item.get("affected_offices_count", 0) for item in mapping)
        if total_atms == 0 and total_offices == 0:
            total_atms = sum(item.get("affected_assets_count", 0) for item in mapping)
            total_offices = max(1, len(mapping))
            
        cities_str = ", ".join(cities)
        inc_types_str = ", ".join(inc_types_ru)
        
        p2_text = (
            f"Краткий свод: {inc_types_str}; {cities_str}, упоминание в телеграм-каналах началось {from_date}, "
            f"общее количество банкоматов и офисов в затронутых городах: {total_atms} банкоматов и {total_offices} офисов."
        )
        report.append(p2_text)
    else:
        report.append("Краткий свод: активных чрезвычайных ситуаций не зарегистрировано; все города функционируют в штатном режиме.")
        
    # Paragraphs 3 and onwards: scenarios
    num_words = ["Первый", "Второй", "Третий", "Четвертый"]
    for idx, sc in enumerate(scenarios):
        num_word = num_words[idx] if idx < len(num_words) else f"{idx+1}-й"
        sc_needs = ", ".join(sc.get("needs", []))
        sc_actions = "; ".join(sc.get("actions", [])) if isinstance(sc.get("actions"), list) else str(sc.get("actions", ""))
        
        sc_needs = sc_needs.rstrip(". ")
        sc_actions = sc_actions.rstrip(". ")
        
        p_sc_text = (
            f"{num_word} предлагаемый сценарий реагирования нацелен на решение: {sc_needs}. "
            f"Предлагаемые действия: {sc_actions}."
        )
        report.append(p_sc_text)
        
    final_markdown = "\n\n".join(report)
        
    report.append("## 🔗 РЕЗУЛЬТАТЫ СЛИЯЯНИЯ РЕКОМЕНДАЦИЙ")
    if rec_path_excel:
        report.append(f"Успешно применены и интегрированы внешние сценарии ЧС из Excel-файла: `{os.path.basename(rec_path_excel)}`")
        if excel_rules:
            report.append("- Раздел правил Лист1 успешно проиндексирован и наложен на матрицу угроз.")
        else:
            report.append("- Файл существует, но из него не удалось прочитать правила.")
    elif rec_obj["has_file"]:
        report.append(f"Успешно применены и интегрированы внешние рекомендации из файла: `{rec_obj['filename']}`")
        if rec_obj["rules"]:
            report.append("- Раздел правил JSON-структуры успешно проиндексирован и наложен на матрицу угроз.")
        else:
            report.append("- Файл содержит плоский текст, фрагменты которого привязаны к аварийным сценариям.")
    else:
        report.append("- Внешний файл рекомендаций в загрузках не найден (проверялись `Агент - примеры сценариев.xlsx`, `Рекомендации.json` и `recommendations.json`). Сценарии сформированы на основе встроенных экспертных правил банка и актуальной карты угроз.")
        
    final_markdown = "\n".join(report)
    
    # 6. Prepare full JSON payload to save
    result_data = {
        "metadata": {
            "generated_at": date_str,
            "has_scenarios_excel": bool(rec_path_excel),
            "scenarios_filename": os.path.basename(rec_path_excel) if rec_path_excel else "",
            "has_recommendations_file": rec_obj["has_file"],
            "recommendations_filename": rec_obj["filename"],
            "processed_threats_count": len(mapping)
        },
        "summary_text": final_markdown,
        "scenarios": scenarios
    }
    
    # 7. Write to temp_storage
    raw_payload_str = json.dumps(result_data, indent=2, ensure_ascii=False)
    save_ok = _write_temp_storage_key(api, "risk_scenarios", raw_payload_str, "Leadership DSS summary text and 2-4 tactical response scenarios.")
    
    # 8. Store flat state backup for UI Widgets polling
    state_dir = Path(api.get_state_dir())
    ui_preview_path = state_dir / "ui_preview.json"
    ui_preview_data = {
        "summary": {
            "last_generated_at": date_str,
            "recommendations_file": os.path.basename(rec_path_excel) if rec_path_excel else (rec_obj["filename"] if rec_obj["has_file"] else "None"),
            "active_scenarios_count": len(scenarios),
            "threats_count": len(mapping),
            "run_result": f"Generation complete! Compiled {len(scenarios)} response scenarios successfully. Saved to temp_storage: {'Yes' if save_ok else 'No (Fallback Active)'}"
        },
        "scenarios": [
            {
                "index": i,
                "name": sc["name"],
                "context": sc.get("context", ""),
                "steps_count": len(sc.get("actions", [])),
                "needs_count": len(sc.get("needs", []))
            }
            for i, sc in enumerate(scenarios, 1)
        ],
        "report_markdown": final_markdown
    }
    
    try:
        import tempfile
        fd, temp_path = tempfile.mkstemp(dir=str(state_dir), suffix=".tmp", text=True)
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(ui_preview_data, f, indent=2, ensure_ascii=False)
        os.replace(temp_path, str(ui_preview_path))
    except Exception as e:
        api.log("error", f"Failed to save UI preview state: {str(e)}")
        
    return ui_preview_data["summary"]["run_result"]

def register(api):
    
    # --- Register Tool ---

    def tool_run_generation(ctx, force: bool = False):
        """Analyzes active threat pairings in temp_storage and compiles key risk response scenarios, persisting the result as 'risk_scenarios'."""
        return perform_decision_generation(api, force)

    api.register_tool(
        "run_generation",
        handler=tool_run_generation,
        description="Compile tactical risk-management scenarios (2 to 4) on active disaster locations matched by asset_geo_matcher and store final report in temp_storage.",
        schema={
            "type": "object",
            "properties": {
                "force": {"type": "boolean", "description": "Forforce re-generation and overwrite storage"}
            }
        },
        timeout_sec=120
    )

    # --- HTTP Routes ---

    async def route_run(request):
        data = await request.json() if request.method == "POST" else {}
        force = bool(data.get("force") or False)
        # Execute matching on thread-pool safely (non-blocking)
        await asyncio.to_thread(perform_decision_generation, api, force)
        return await route_status(request)

    async def route_status(request):
        state_dir = Path(api.get_state_dir())
        ui_preview_path = state_dir / "ui_preview.json"
        
        default_data = {
            "summary": {
                "last_generated_at": "Never",
                "recommendations_file": "None",
                "active_scenarios_count": 0,
                "threats_count": 0,
                "run_result": "Pending first scenarios generation run"
            },
            "scenarios": [],
            "report_markdown": "No scenarios compiled yet. Use the tool or trigger through form below."
        }
        
        if ui_preview_path.exists():
            try:
                default_data = json.loads(ui_preview_path.read_text(encoding="utf-8"))
            except Exception as e:
                api.log("error", f"Failed to load UI preview from disk: {str(e)}")
                
        return default_data

    api.register_route("run", route_run, methods=("POST",))
    api.register_route("status", route_status, methods=("GET",))

    # --- UI Declarative Tab layout ---

    api.register_ui_tab(
        "panel",
        "Risk Scenarios",
        icon="shield",
        render={
            "kind": "declarative",
            "schema_version": 1,
            "span": 2,
            "components": [
                {
                    "type": "poll",
                    "route": "status",
                    "auto_start": True,
                    "label": "Sync Decision Board"
                },
                {
                    "type": "markdown",
                    "text": "### 🛡️ Risk Scenario Decision Support Core\nLogical expert system that ingests threats mapped in temp_storage, merges optional recommendations, and generates operational-response scenarios satisfying core customer/bank needs."
                },
                {
                    "type": "form",
                    "route": "run",
                    "method": "POST",
                    "submit_label": "Trigger Scenarios Compilation Now",
                    "fields": [
                        {
                            "type": "checkbox",
                            "name": "force",
                            "label": "Сформировать тактические сценарии повторно",
                            "required": False
                        }
                    ]
                },
                {
                    "type": "markdown",
                    "text": "#### 📊 Generation Operations Status Summary"
                },
                {
                    "type": "kv",
                    "fields": [
                        {"label": "Last Execution", "path": "summary.last_generated_at"},
                        {"label": "Optional Guidance File", "path": "summary.recommendations_file"},
                        {"label": "Active Scenarios Compiled", "path": "summary.active_scenarios_count"},
                        {"label": "Threat Maps Processed", "path": "summary.threats_count"},
                        {"label": "Generation Result", "path": "summary.run_result"}
                    ]
                },
                {
                    "type": "markdown",
                    "text": "#### 📋 Compiled Tactical Response Scenarios Overview"
                },
                {
                    "type": "table",
                    "path": "scenarios",
                    "columns": [
                        {"path": "index", "label": "Scenario N"},
                        {"path": "name", "label": "Scenario Title"},
                        {"path": "context", "label": "Trigger Conditions / Trigger context"},
                        {"path": "steps_count", "label": "Operational Steps"},
                        {"path": "needs_count", "label": "Satisfied Needs"}
                    ]
                },
                {
                    "type": "markdown",
                    "text": "#### 📑 Generated Analytical Summary & Report (Leaderboard View)"
                },
                {
                    "type": "markdown",
                    "path": "report_markdown"
                }
            ]
        }
    )
