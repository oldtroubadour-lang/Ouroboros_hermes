import json
import os
import urllib.request
import asyncio
from pathlib import Path
from datetime import datetime

def _get_server_port(api) -> str:
    state_dir = Path(api.get_state_dir())
    data_root = state_dir.parent.parent / "server_port"
    port = "8765"
    if data_root.exists():
        try:
            port = data_root.read_text(encoding="utf-8").strip()
        except Exception:
            pass
    return port

def _read_temp_storage_key(api, key: str) -> str:
    port = _get_server_port(api)
    url = f"http://127.0.0.1:{port}/api/extensions/temp_storage/retrieve"
    payload = {"key": key}
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            res = json.loads(r.read().decode("utf-8"))
            retrieved = res.get("retrieved") or {}
            if retrieved.get("key") == key:
                return retrieved.get("value") or ""
    except Exception as e:
        api.log("warning", f"HTTP POST to temp_storage retrieve failed: {str(e)}")
        
    # Fully conformed, secure local state folder fallback
    state_dir = Path(api.get_state_dir())
    local_fallback = state_dir / f"fallback_{key}.json"
    if local_fallback.exists():
        try:
            return local_fallback.read_text(encoding="utf-8")
        except Exception as exc:
            api.log("error", f"Local fallback read failed: {str(exc)}")
            
    return ""

def _write_temp_storage_key(api, key: str, value: str, description: str = "") -> bool:
    port = _get_server_port(api)
    url = f"http://127.0.0.1:{port}/api/extensions/temp_storage/store"
    payload = {
        "key": key,
        "value": value,
        "description": description
    }
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            res = json.loads(r.read().decode("utf-8"))
            if res and "error" not in res:
                return True
    except Exception as e:
        api.log("warning", f"HTTP POST to temp_storage store failed: {str(e)}")
        
    # Safe local fallback write strictly confined inside our own state directory (Zero sibling write crossing)
    try:
        state_dir = Path(api.get_state_dir())
        state_dir.mkdir(parents=True, exist_ok=True)
        local_fallback = state_dir / f"fallback_{key}.json"
        local_fallback.write_text(value, encoding="utf-8")
    except Exception as exc:
        api.log("error", f"Local fallback write failed: {str(exc)}")
    return False

def check_temp_storage_health(api) -> bool:
    port = _get_server_port(api)
    url = f"http://127.0.0.1:{port}/api/extensions/temp_storage/status"
    try:
        req = urllib.request.Request(url, method="GET")
        with urllib.request.urlopen(req, timeout=3) as r:
            return r.status == 200
    except Exception:
        return False

def find_excel_registry(api) -> str:
    user_home = os.environ.get('USERPROFILE', os.path.expanduser('~'))
    candidates = [
        os.path.join(user_home, 'Downloads', 'Справочник объектов для агента.xlsx'),
        os.path.join(user_home, 'Downloads', 'Справочник объектов для агента-1.xlsx'),
        os.path.join(user_home, 'Downloads', 'Справочник объектов для агента-2.xlsx'),
        os.path.join(api.get_state_dir(), 'Справочник объектов для агента.xlsx'),
        os.path.join(user_home, 'Downloads', 'Данные для агента.xlsx'),  # fallback
        os.path.join(api.get_state_dir(), 'Данные для агента.xlsx')      # fallback
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return ""

def match_assets_and_threats(api, force: bool = False) -> str:
    api.log("info", "Starting asset_geo_matcher calculations.")
    
    # 1. Preflight dependency and tools import check
    try:
        import openpyxl
    except ImportError:
        api.log("error", "openpyxl is not installed in the python environment.")
        return "Error: Python dependency 'openpyxl' is not installed."
        
    excel_path = find_excel_registry(api)
    if not excel_path:
        api.log("error", "Excel registry file 'Справочник объектов для агента.xlsx' not found in Downloads or state.")
        return "Error: Excel registry file 'Справочник объектов для агента.xlsx' was not found in Downloads directory."
        
    temp_storage_healthy = check_temp_storage_health(api)
    if not temp_storage_healthy:
        api.log("warning", "temp_storage extension is not reachable. Matcher will write locally and save fallbacks.")
        
    # 2. Get active incidents from temp_storage
    raw_incidents = _read_temp_storage_key(api, "active_incidents")
    if not raw_incidents:
        api.log("warning", "No active incidents found in temp_storage key 'active_incidents'.")
        return "Error: No active incidents found under 'active_incidents' key in temp_storage."
        
    try:
        incidents = json.loads(raw_incidents)
    except Exception as e:
        api.log("error", f"Failed to parse active_incidents JSON: {str(e)}")
        return f"Error: Active incidents JSON parsing failure: {str(e)}"
        
    if not incidents:
        return "Warning: Active incidents list is empty. No matching can be performed."
        
    # 3. Load Excel workbook and find target sheet
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
    except Exception as e:
        api.log("error", f"Failed to open Excel workbook: {str(e)}")
        return f"Error: Failed to load Excel workbook: {str(e)}"
        
    sheet_name = "Справочник объектов"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
        api.log("warning", f"Sheet 'Справочник объектов' not found, falling back to first sheet: {sheet_name}")
        
    ws = wb[sheet_name]
    
    # Validate required columns/headers from first row schema (preflight check)
    header_row = next(ws.iter_rows(max_row=1, values_only=True), None)
    if not header_row or len(header_row) < 3:
        wb.close()
        api.log("error", f"Invalid sheet schema in excel file at sheet: {sheet_name}")
        return "Error: Invalid Excel sheet schema. The sheet must contain at least Object ID, Asset Type, and Address columns."
        
    mapping = []
    total_matched_assets = 0
    threatened_list_for_ui = [] # Flat list of assets affected for UI preview
    
    # 4. Perform geospatial city mapping
    for inc in incidents:
        city = inc.get("location", {}).get("city")
        if not city or city in ["Неизвестно", "Unknown", ""]:
            continue
            
        affected = []
        city_lower = city.lower()
        
        # Stream rows efficiently using read_only
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                continue # Skip header
                
            if len(row) < 3:
                continue
                
            obj_id, obj_type, address = row[:3]
            kic_code = row[3] if len(row) > 3 else ""
            kic_name = row[4] if len(row) > 4 else ""
            
            if not address:
                continue
                
            address_str = str(address).lower()
            
            # Substring comparison for robust format mapping
            if city_lower in address_str:
                asset_record = {
                    "id": str(obj_id),
                    "type": str(obj_type),
                    "address": str(address),
                    "kic_code": str(kic_code) if kic_code is not None else "",
                    "kic_name": str(kic_name) if kic_name is not None else ""
                }
                affected.append(asset_record)
                total_matched_assets += 1
                
                # Flat preview list for UI
                if len(threatened_list_for_ui) < 100:
                    threatened_list_for_ui.append({
                        "id": str(obj_id),
                        "type": str(obj_type),
                        "city": city,
                        "incident_type": inc.get("incident_type_ru", inc.get("incident_type", "Emergency")),
                        "address": str(address)
                    })
                    
        mapping.append({
            "incident_id": inc.get("incident_id", "unknown"),
            "incident_type": inc.get("incident_type"),
            "incident_type_ru": inc.get("incident_type_ru", "Чрезвычайная ситуация"),
            "city": city,
            "severity": inc.get("severity", "low"),
            "affected_assets_count": len(affected),
            "affected_assets": affected
        })
        
    wb.close()
    
    # 5. Persist mapping
    mapping_str = json.dumps(mapping, indent=2, ensure_ascii=False)
    save_ok = _write_temp_storage_key(api, "threat_mapping", mapping_str, "Geospatial threat pairing matching bank assets with current emergencies.")
    
    # Build complete execution details
    last_matched_str = datetime.utcnow().isoformat() + "Z"
    save_status = "Output successfully saved to temp_storage." if save_ok else "Saved to local fallback (temp_storage was unreachable)."
    run_msg = f"Geospatial threat matching completed! Processed: {len(mapping)} active threat cities. Paired: {total_matched_assets} bank branches or ATMs. {save_status}"
    
    # Save a flat preview to our own state directory to feed UI widget polls
    state_dir = Path(api.get_state_dir())
    ui_preview_path = state_dir / "ui_preview.json"
    ui_preview_data = {
        "summary": {
            "last_matched_at": last_matched_str,
            "total_incidents_processed": len(mapping),
            "total_assets_threatened": total_matched_assets,
            "excel_source": os.path.basename(excel_path),
            "run_result": run_msg
        },
        "incidents_summary": [
            {
                "city": m["city"],
                "incident_type": m["incident_type_ru"],
                "severity": m["severity"],
                "count": m["affected_assets_count"]
            }
            for m in mapping
        ],
        "threatened_assets": threatened_list_for_ui
    }
    
    try:
        import tempfile
        fd, temp_path = tempfile.mkstemp(dir=str(state_dir), suffix=".tmp", text=True)
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(ui_preview_data, f, indent=2, ensure_ascii=False)
        os.replace(temp_path, str(ui_preview_path))
    except Exception as e:
        api.log("error", f"Failed to save UI preview JSON: {str(e)}")
        
    return run_msg

def register(api):
    
    # --- Tool Definition ---

    def tool_run_match(ctx, force: bool = False):
        """Runs the mapping job, scanning open incidents and pairing them with associated ATMs/branches from Downloads registry, updating temp_storage 'threat_mapping'."""
        return match_assets_and_threats(api, force)

    api.register_tool(
        "run_match",
        handler=tool_run_match,
        description="Pair active bank branches and ATMs with current emergency incident cities and store threat mapping in temp_storage.",
        schema={
            "type": "object",
            "properties": {
                "force": {"type": "boolean", "description": "Ensure a direct re-scan of files immediately"}
            }
        },
        timeout_sec=120
    )

    # --- HTTP Route Handlers ---

    async def route_run(request):
        data = await request.json() if request.method == "POST" else {}
        force = bool(data.get("force") or False)
        # Process matching synchronously within thread-pool
        await asyncio.to_thread(match_assets_and_threats, api, force)
        return await route_status(request)

    async def route_status(request):
        state_dir = Path(api.get_state_dir())
        ui_preview_path = state_dir / "ui_preview.json"
        
        preview_data = {
            "summary": {
                "last_matched_at": "Never",
                "total_incidents_processed": 0,
                "total_assets_threatened": 0,
                "excel_source": "None",
                "run_result": "Pending first match run"
            },
            "incidents_summary": [],
            "threatened_assets": []
        }
        
        if ui_preview_path.exists():
            try:
                preview_data = json.loads(ui_preview_path.read_text(encoding="utf-8"))
            except Exception as e:
                api.log("error", f"Failed to load UI preview from disk: {str(e)}")
                
        return preview_data

    api.register_route("run", route_run, methods=("POST",))
    api.register_route("status", route_status, methods=("GET",))

    # --- UI Declarative Tab layout ---

    api.register_ui_tab(
        "panel",
        "Asset Matcher",
        icon="map",
        render={
            "kind": "declarative",
            "schema_version": 1,
            "span": 2,
            "components": [
                {
                    "type": "poll",
                    "route": "status",
                    "auto_start": True,
                    "label": "Sync Matcher Board"
                },
                {
                    "type": "markdown",
                    "text": "### 🗺️ Physical Bank Assets & Threats Geospatial Matcher\nPairs bank offices and ATMs with active disaster zones extracted from Telegram channels."
                },
                {
                    "type": "form",
                    "route": "run",
                    "method": "POST",
                    "submit_label": "Trigger Geospatial Matching Run Now",
                    "fields": [
                        {
                            "type": "checkbox",
                            "name": "force",
                            "label": "Принудительно сопоставить данные заново",
                            "required": False
                        }
                    ]
                },
                {
                    "type": "markdown",
                    "text": "#### 📊 Matcher Operations Status Summary"
                },
                {
                    "type": "kv",
                    "fields": [
                        {"label": "Last Execution", "path": "summary.last_matched_at"},
                        {"label": "Source Excel Database", "path": "summary.excel_source"},
                        {"label": "Active Cities Scanned", "path": "summary.total_incidents_processed"},
                        {"label": "Total Bank Assets in Danger Zone", "path": "summary.total_assets_threatened"},
                        {"label": "Match Execution Result", "path": "summary.run_result"}
                    ]
                },
                {
                    "type": "markdown",
                    "text": "#### ⚠️ Threat Levels by City"
                },
                {
                    "type": "table",
                    "path": "incidents_summary",
                    "columns": [
                        {"path": "city", "label": "Incident City"},
                        {"path": "incident_type", "label": "Type of Threat"},
                        {"path": "severity", "label": "Severity"},
                        {"path": "count", "label": "Threatened Bank Assets"}
                    ]
                },
                {
                    "type": "markdown",
                    "text": "#### 📋 Threatened Bank Branches & ATMs List (Preview)"
                },
                {
                    "type": "table",
                    "path": "threatened_assets",
                    "columns": [
                        {"path": "id", "label": "Object ID"},
                        {"path": "type", "label": "Asset Type"},
                        {"path": "city", "label": "City"},
                        {"path": "incident_type", "label": "Associated Danger"},
                        {"path": "address", "label": "Asset Address"}
                    ]
                }
            ]
        }
    )
